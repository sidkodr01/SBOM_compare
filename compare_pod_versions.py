import sys
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def read_user_data(file_path):
    ext = os.path.splitext(file_path)[1].lower()
    if ext == '.txt':
        df = pd.read_csv(file_path)
    elif ext == '.xlsx':
        df = pd.read_excel(file_path)
    else:
        raise ValueError('Unsupported user data file type. Use .txt or .xlsx')
    return df


def read_expected_versions(file_path):
    return pd.read_excel(file_path)


def compare_versions(user_df, expected_df, pod_col='pod', version_col='version'):
    # Merge on pod name
    merged = pd.merge(user_df, expected_df, on=pod_col, how='left', suffixes=('', '_expected'))
    # Add a column for mismatch
    merged['mismatch'] = merged[version_col] != merged[version_col + '_expected']
    return merged


def highlight_mismatches(output_path, merged_df, pod_col='pod', version_col='version'):
    # Write to Excel first
    merged_df.to_excel(output_path, index=False)
    # Load with openpyxl
    wb = load_workbook(output_path)
    ws = wb.active
    # Find columns
    version_idx = None
    mismatch_idx = None
    expected_idx = None
    for idx, cell in enumerate(ws[1], 1):
        if cell.value == version_col:
            version_idx = idx
        elif cell.value == 'mismatch':
            mismatch_idx = idx
        elif cell.value == version_col + '_expected':
            expected_idx = idx
    # Check that all required columns are found
    if version_idx is None or mismatch_idx is None or expected_idx is None:
        raise ValueError('Could not find required columns in the output Excel file.')
    # Red fill
    red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    # Highlight mismatches
    for row in range(2, ws.max_row + 1):
        mismatch_cell = ws.cell(row, mismatch_idx)
        if mismatch_cell.value:
            ws.cell(row, version_idx).fill = red_fill
            ws.cell(row, expected_idx).fill = red_fill
    # Save
    wb.save(output_path)


def main():
    if len(sys.argv) != 4:
        print('Usage: python compare_pod_versions.py <user_data_file> <expected_versions_file> <output_file>')
        sys.exit(1)
    user_data_file = sys.argv[1]
    expected_versions_file = sys.argv[2]
    output_file = sys.argv[3]

    user_df = read_user_data(user_data_file)
    expected_df = read_expected_versions(expected_versions_file)

    # Assume columns are named 'pod' and 'version' in both files
    merged_df = compare_versions(user_df, expected_df, pod_col='pod', version_col='version')
    highlight_mismatches(output_file, merged_df, pod_col='pod', version_col='version')
    print(f'Output written to {output_file}')


if __name__ == '__main__':
    main() 