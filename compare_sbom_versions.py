import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# File paths
actual_file = 'SBOM Summary Fixed.xlsx'
expected_file = 'sbom_summary_fixed - Copy_sid_changed.xlsx'
output_file = 'SBOM_Version_Comparison_Result.xlsx'
temp_file = 'temp_combined.xlsx'

# Load Excel files
df_actual = pd.read_excel(actual_file)
df_expected = pd.read_excel(expected_file)

# Columns to compare
version_columns = [
    'Detected Base Image',
    'Detected Tomcat Version',
    'Detected Spring Boot Version',
    'Detected Java Version'
]

# Add expected columns
for col in version_columns:
    expected_col = f'Expected {col.split("Detected ")[-1]}'
    df_actual[expected_col] = df_expected[col]

# Compute STATUS column
def compute_status(row):
    for col in version_columns:
        expected_col = f'Expected {col.split("Detected ")[-1]}'
        actual_val = str(row[col]).strip() if pd.notna(row[col]) else ''
        expected_val = str(row[expected_col]).strip() if pd.notna(row[expected_col]) else ''
        if actual_val != expected_val:
            return 'MISMATCH'
    return 'MATCH'

df_actual['STATUS'] = df_actual.apply(compute_status, axis=1)

# Save combined DataFrame to Excel
df_actual.to_excel(temp_file, index=False)

# Apply red fill using openpyxl
wb = load_workbook(temp_file)
ws = wb.active
red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

# Header mapping
header = [cell.value for cell in ws[1]]
col_index = {col: idx + 1 for idx, col in enumerate(header)}

# Highlight mismatches
for row in range(2, ws.max_row + 1):
    for col in version_columns:
        actual_cell = ws.cell(row=row, column=col_index[col])
        expected_col = f'Expected {col.split("Detected ")[-1]}'
        expected_cell = ws.cell(row=row, column=col_index[expected_col])

        actual_val = str(actual_cell.value).strip() if actual_cell.value else ''
        expected_val = str(expected_cell.value).strip() if expected_cell.value else ''

        if actual_val != expected_val:
            actual_cell.fill = red_fill
            expected_cell.fill = red_fill

# Highlight STATUS column if mismatch
status_col_idx = col_index['STATUS']
for row in range(2, ws.max_row + 1):
    status_cell = ws.cell(row=row, column=status_col_idx)
    if status_cell.value == 'MISMATCH':
        status_cell.fill = red_fill

# Save final output
wb.save(output_file)
print(f"✅ Comparison complete. Output saved to: {output_file}")
