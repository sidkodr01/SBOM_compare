import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# File paths
input_file = "SBOM_Version_Comparison_Result.xlsx"
output_file = "sbom_comparison_with_summary1.xlsx"

# Load Excel data
df = pd.read_excel(input_file)

# Utility to check if cell has value
def has_value(x):
    return pd.notna(x) and str(x).strip() != ""

# Define helper flags
df['Is_UBI'] = df['Detected Base Image'].str.lower().str.contains('ubi', na=False)
df['Is_Java'] = df['Detected Java Version'].apply(has_value)
df['Has_Tomcat'] = df['Detected Tomcat Version'].apply(has_value)
df['Has_SpringBoot'] = df['Detected Spring Boot Version'].apply(has_value)

df['Java_With_Tomcat_Spring'] = df['Is_Java'] & (df['Has_Tomcat'] | df['Has_SpringBoot'])
df['Java_Without_Tomcat_Spring'] = df['Is_Java'] & ~(df['Has_Tomcat'] | df['Has_SpringBoot'])
df['Non_Java'] = ~df['Is_Java']

df['Java_With_Tomcat_Spring_OK'] = df['Java_With_Tomcat_Spring'] & (df['STATUS'] == 'MATCH')
df['Java_Without_Tomcat_Spring_OK'] = df['Java_Without_Tomcat_Spring'] & (df['STATUS'] == 'MATCH')
df['Non_Java_OK'] = df['Non_Java'] & (df['STATUS'] == 'MATCH')

# Create Anomaly Reason column
df['Anomaly Reason'] = ''
mask = (df['STATUS'] == 'MISMATCH') & (~df['Has_Tomcat']) & (~df['Has_SpringBoot'])
df.loc[mask, 'Anomaly Reason'] = 'MISMATCH and missing Spring Boot/Tomcat'

# Create summary
summary_data = [
    ['Total Images', len(df)],
    ['UBI Migrated Images', df['Is_UBI'].sum()],
    ['UBI "Lineup OK" Images', df[df['Is_UBI'] & (df['STATUS'] == 'MATCH')].shape[0]],
    ['Java images with Tomcat/SpringBoot "lineup OK"', df['Java_With_Tomcat_Spring_OK'].sum()],
    ['Total Java images with Tomcat/SpringBoot', df['Java_With_Tomcat_Spring'].sum()],
    ['Java images without Tomcat/SpringBoot "lineup OK"', df['Java_Without_Tomcat_Spring_OK'].sum()],
    ['Non-Java images "lineup OK"', df['Non_Java_OK'].sum()],
    ['Images with MISMATCH and missing Spring Boot/Tomcat', df['Anomaly Reason'].eq('MISMATCH and missing Spring Boot/Tomcat').sum()],
]

summary_df = pd.DataFrame(summary_data, columns=["Metric", "Value"])

# Step 1: Save both sheets
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='SBOM Comparison', index=False)
    summary_df.to_excel(writer, sheet_name='Summary', index=False)

# Step 2: Reopen and highlight anomalies
wb = load_workbook(output_file)
ws = wb['SBOM Comparison']

# Red fill style
red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

# Get the column index of "Anomaly Reason"
header_row = [cell.value for cell in ws[1]]
if 'Anomaly Reason' not in header_row:
    raise ValueError("Anomaly Reason column not found!")

anomaly_col_index = header_row.index('Anomaly Reason') + 1  # Excel is 1-indexed

# Step 3: Iterate through rows and highlight if anomaly exists
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    reason_cell = row[anomaly_col_index - 1]
    if reason_cell.value == 'MISMATCH and missing Spring Boot/Tomcat':
        for cell in row:
            cell.fill = red_fill

# Step 4: Save final file
wb.save(output_file)
print(f"✅ File saved with working row highlights: {output_file}")
